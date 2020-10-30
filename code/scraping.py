import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

def get_page(url):
    return requests.get(url).content


def get_paper_info(url):
    """Scrape abstract and keywords from given Springer-Link paper url"""
    results = {}
    if len(url) == 0:
        return None
    page = get_page(url)
    soup = BeautifulSoup(page, features="html.parser")

    section = soup.find(id="Abs1")
    paragraphs = section.find_all("p")
    ps = []
    for p in paragraphs:
        ps.append(p.text)
    abstract = " ".join(ps)

    keyword_elements = soup.find_all('span',class_="Keyword")
    keywords = []
    if len(keyword_elements) != 0:
        for word in keyword_elements:
            try:
                keywords.append(word.string.strip())
            except AttributeError as e:
                continue

        keywords = ", ".join(keywords)

    else:
        keywords = ""

    results['abstract'] = abstract
    results['keywords'] = keywords
    return results

def add_abstracts_and_keywords(path='../gradu/material/ICAGI.xlsx', target="../gradu/material/articles_combined.xlsx"):
    """Add info from given articles to material_search file"""
    wb = load_workbook(path)
    ws = wb.active
    wb_new = load_workbook("../gradu/material_search.xlsx")
    new_sheet = wb_new.active
    number_of_articles_to_be_added = 152
    source_start_row = 2
    target_row = 39
    number_col = 1
    title_col = 2
    author_col = 3
    year_col = 4
    venue_col = 5
    abstract_col = 6
    keywords_col = 7
    link_col = 8
    doi_col = 9
    cite_col = 10
    type_col = 11
    access_col = 12
    source_col = 13
    volume_col = 14

    for row in range(source_start_row, source_start_row + number_of_articles_to_be_added):
        title = ws.cell(row=row, column=1).value
        doi = ws.cell(row=row, column=6).value
        authors = ws.cell(row=row, column=7).value
        year = ws.cell(row=row, column=8).value
        url = ws.cell(row=row, column=9).value

        abstract = get_paper_info(url)['abstract']
        keywords = get_paper_info(url)['keywords']

        new_sheet.cell(row=target_row, column=abstract_col).value = abstract
        new_sheet.cell(row=target_row, column=keywords_col).value = keywords
        new_sheet.cell(row=target_row, column=title_col).value = title
        new_sheet.cell(row=target_row, column=doi_col).value = doi
        new_sheet.cell(row=target_row, column=author_col).value = authors
        new_sheet.cell(row=target_row, column=year_col).value = year
        
        new_sheet.cell(row=target_row, column=type_col).value = "Conference Paper"
        new_sheet.cell(row=target_row, column=venue_col).value = "ICAGI"
        new_sheet.cell(row=target_row, column=link_col).value = url
        new_sheet.cell(row=target_row, column=source_col).value = "Springer-Link"
        target_row += 1
        
    wb_new.save(target)

if __name__ == "__main__":
    add_abstracts_and_keywords()