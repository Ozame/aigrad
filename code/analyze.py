import pprint
import sys
import csv
from collections import OrderedDict
from dataclasses import asdict, dataclass, field
from typing import Dict, List

import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from nltk.probability import FreqDist
from openpyxl import load_workbook
from openpyxl.styles import Font
from pandas.core.frame import DataFrame
from geopy.geocoders import Nominatim

# Wieringa classes: ER, VR, SP, PP, OP, PEP
W_CLASSES = [
    ("er", "Evaluation research"),
    ("vr", "Validation research"),
    ("sp", "Solution proposal"),
    ("pp", "Philosophical paper"),
    ("op", "Opinion paper"),
    ("pep", "Personal experience paper"),
]

CATS = [
    (1, "Cognitive architectures"),
    (2, "AGI design"),
    (3, "Reasoning and Inference"),
    (4, "Planning and decision making"),
    (5, "Probabilistic approaches"),
    (6, "Category theory"),
    (7, "Universal AI"),
    (8, "Physical robots"),
    (9, "Computer vision and perception"),
    (10, "Nature-inspired approaches"),
    (11, "Reinforcement learning"),
    (12, "Recursive self-Improvement"),
    (13, "Experiential learning"),
    (14, "Agent environment"),
    (15, "Multi-agent systems"),
    (16, "Human-computer interaction"),
    (17, "AI safety"),
    (18, "Philosophical aspects"),
    (19, "Human-like qualities"),
    (20, "AGI research"),
    (21, "AI evaluation"),
    (22, "Game playing"),
]


@dataclass
class Paper:
    number: int
    name: str
    keywords: List[str]
    w_classes: List[str]
    categories: List[str] = field(default_factory=list)
    countries: List[str] = field(default_factory=list)
    authors: List[str] = field(default_factory=list)
    year: int = None
    venue: str = None
    abstract: str = None


def parse_cat_numbers(s):
    if type(s) == int:
        return [s]
    cats = filter(lambda x: len(x.strip()), s.split(sep=","))
    return list(map(int, cats))


def parse_words(s: str):
    words = s.split(sep=",")
    return list(
        filter(
            lambda x: 1 < len(x),
            map(lambda x: x.lower().strip().replace("(", "").replace(")", ""), words),
        )
    )


def load_papers(path: str = "../gradu/material/final_results.xlsx") -> List[Paper]:
    """Loads the papers and returns their relevant information as list of Papers"""
    number_col = "A"
    name_col = "B"
    keywords_col = "M"
    w_classes_col = "N"
    year_col = "D"
    venue_col = "E"
    abstract_col = "F"
    cat_column = "Q"
    countries_col = "S"
    authors_col = "C"
    papers = []

    wb = load_workbook(path)
    sheet = wb["Accepted Papers"]
    for i in range(3, 95):
        row = str(i)
        paper = Paper(
            number=sheet[number_col + row].value,
            name=sheet[name_col + row].value,
            keywords=parse_words(sheet[keywords_col + row].value),
            w_classes=parse_words(sheet[w_classes_col + row].value),
            categories=parse_cat_numbers(sheet[cat_column + row].value),
            countries=parse_words(sheet[countries_col + row].value),
            authors=parse_words(sheet[authors_col + row].value),
            year=sheet[year_col + row].value,
            venue=sheet[venue_col + row].value,
            abstract=sheet[abstract_col + row].value,
        )
        papers.append(paper)
    return papers


def load_papers_df(path: str = "../gradu/material/final_results.xlsx"):
    """Loads the paper data as a pandas dataframe"""
    papers = load_papers(path)
    df = DataFrame(
        columns=["number", "name", "year", "venue", "er", "vr", "sp", "pp", "op", "pep"]
        + list(map(lambda x: "c" + str(x), range(1, 23)))
    )
    # Encodes classes and categories into separate columns
    for paper in papers:
        p = asdict(paper)
        w_classes = p["w_classes"]
        cats = p["categories"]
        for c in (x[0] for x in W_CLASSES):
            if c in w_classes:
                p[c] = 1
            else:
                p[c] = 0
        for cat in range(1, 23):
            if cat in cats:
                p["c" + str(cat)] = 1
            else:
                p["c" + str(cat)] = 0

        # Deletes unnecessary keys
        del p["abstract"]
        del p["keywords"]
        del p["w_classes"]
        del p["categories"]
        df = df.append(pd.Series(p), ignore_index=True)
    return df


def create_article_csv(
    papers: List[Paper], path="../gradu/material/data/accepted_papers.csv"):
    
    header = "year,authors,title,class,categories".split(sep=",")
    rows = []
    
    for paper in papers:
        authors = ", ".join(map(lambda x: x.title(), paper.authors))
        classes = ", ".join(map(lambda x: dict(W_CLASSES)[x], paper.w_classes))
        categories = ", ".join(map(lambda x: dict(CATS)[x], paper.categories))
        row = [paper.year, authors, paper.name, classes, categories]
        rows.append(row)
    rows = sorted(rows, key=lambda x: x[0])

    with open(path, "w", newline="", encoding="utf-8") as f:
        csvwriter = csv.writer(f, delimiter=";")
        csvwriter.writerow(header)
        csvwriter.writerows(rows)


def load_categories(
    path: str = "../gradu/material/final_results.xlsx",
) -> Dict[str, str]:
    """
    Loads the categories and their respective keywords

    Returns a dict with keywords as keys and their respective categories as values
    """
    categorization = {}
    wb = load_workbook(path)
    if not "Categorization" in wb:
        return None
    sheet = wb["Categorization"]
    for row in sheet.iter_rows(min_row=3, max_row=297, min_col=2, max_col=15):
        for cell in row:
            if cell.value and "/" in cell.value:
                keyword = cell.value[: cell.value.rfind("/")]
                category = sheet["A" + str(cell.row)].value or None
                categorization[keyword] = category
    return categorization


def count_keywords(path: str = "../gradu/material/final_results.xlsx"):
    """Counts the occurrences of keyword instances on the same row"""
    result_column = "Q"
    results_row = 3

    wb = load_workbook(path)
    sheet = wb["Categorization"]
    for row in sheet.iter_rows(min_row=results_row, max_row=297, min_col=2, max_col=15):
        count = 0

        for cell in row:
            if cell.value and "/" in cell.value:
                count += int(cell.value[cell.value.rfind("/") + 1 :])
            results_row = cell.row
        result_cell = sheet[result_column + str(results_row)]
        result_cell.value = count
        if 4 <= count:
            result_cell.font = Font(bold=True)

    wb.save(path)


def calculate_category_distribution(path: str = "../gradu/material/final_results.xlsx"):
    """Calculates how the articles would be divided based on the current categorization"""
    papers = load_papers()
    categories = load_categories()
    dist = {}
    empty_papers = []
    # Every possible category for paper is fetched
    for paper in papers:
        cats = []
        for kw in paper.keywords:
            cat = categories[kw]
            if cat not in cats:
                cats.append(cat)
            else:
                continue
        # If multiple categories are found, None-category is removed
        if 1 < len(cats) and None in cats:
            cats = [i for i in cats if i]
        # paper's final categories are added to the distribution
        # Notice that paper.categories is not assigned as it is unnecessary
        if len(cats) == 1 and None in cats:
            empty_papers.append(paper)
        for cat in cats:
            if cat in dist:
                dist[cat] += 1
            else:
                dist[cat] = 1

    pp = pprint.PrettyPrinter()
    c = sum(dist.values())
    print("Total categorizations: " + str(c))
    pp.pprint(dist)
    print("The following are papers with no categories:")
    pp.pprint(empty_papers)
    return dist


def write_category_dist_to_workbook(
    c_dist: Dict[str, int], path: str = "../gradu/material/final_results.xlsx"
):
    """Writes the given category distribution to the worksheet"""

    row_number = 6
    category_name_column = "T"
    category_count_column = "U"

    wb = load_workbook(path)
    if not "Categorization" in wb:
        return
    sheet = wb["Categorization"]
    for name, count in sorted(c_dist.items(), key=lambda x: x[1], reverse=True):
        sheet[category_name_column + str(row_number)] = name
        sheet[category_count_column + str(row_number)] = count
        row_number += 2

    wb.save(path)


def calculate_category_count(
    path: str = "../gradu/material/final_results.xlsx",
) -> Dict[int, int]:
    """Calculates the number of papers in each category"""
    number_col = "A"
    name_col = "B"
    keywords_col = "M"
    w_classes_col = "N"
    category_column = "Q"
    dist = dict([(x, 0) for x in range(1, 23)])

    wb = load_workbook(path)
    sheet = wb["Accepted Papers"]
    for i in range(3, 95):
        row = str(i)
        cat_string = sheet[category_column + row].value
        cats = parse_cat_numbers(cat_string)
        for cat in cats:
            dist[cat] += 1

    sorted_dist = OrderedDict(sorted(dist.items(), key=lambda x: x[1], reverse=True))
    return sorted_dist


def get_category_papers(
    cat_numbers: List[int],
    path: str = "../gradu/material/final_results.xlsx",
) -> Dict[int, int]:
    """Returns the list of papers in given categories. Paper must belong to every given category"""
    number_col = "A"
    name_col = "B"
    keywords_col = "M"
    w_classes_col = "N"
    category_column = "Q"
    countries_col = "S"
    papers = []

    if not cat_numbers:
        return None

    wb = load_workbook(path)
    sheet = wb["Accepted Papers"]
    for i in range(3, 95):
        row = str(i)
        cat_string = sheet[category_column + row].value
        cats = parse_cat_numbers(cat_string)
        match_counter = 0
        for number in cat_numbers:
            if number in cats:
                match_counter += 1
        if match_counter == len(cat_numbers):
            papers.append(
                Paper(
                    number=sheet[number_col + row].value,
                    name=sheet[name_col + row].value,
                    keywords=parse_words(sheet[keywords_col + row].value),
                    w_classes=parse_words(sheet[w_classes_col + row].value),
                    categories=list((map(str, cats))),
                    countries=parse_words(sheet[countries_col + row].value),
                )
            )

    sorted_list = sorted(papers, key=lambda x: x.number, reverse=True)
    return sorted_list


def get_country_papers(
    countries: List[str],
    path: str = "../gradu/material/final_results.xlsx",
) -> Dict[int, int]:
    """Returns the list of papers in given countries. Paper must belong to every given country"""
    number_col = "A"
    name_col = "B"
    keywords_col = "M"
    w_classes_col = "N"
    category_column = "Q"
    countries_col = "S"
    authors_col = "C"
    papers = []

    if not countries:
        return None

    wb = load_workbook(path)
    sheet = wb["Accepted Papers"]
    for i in range(3, 95):
        row = str(i)
        country_string = sheet[countries_col + row].value
        paper_countries = parse_words(country_string)
        match_counter = 0
        for number in countries:
            if number in paper_countries:
                match_counter += 1
        if match_counter == len(countries):
            papers.append(
                Paper(
                    number=sheet[number_col + row].value,
                    name=sheet[name_col + row].value,
                    keywords=parse_words(sheet[keywords_col + row].value),
                    w_classes=parse_words(sheet[w_classes_col + row].value),
                    categories=parse_cat_numbers(sheet[category_column + row].value),
                    countries=parse_words(sheet[countries_col + row].value),
                    authors=parse_words(sheet[authors_col + row].value),
                )
            )

    sorted_list = sorted(papers, key=lambda x: x.number, reverse=True)
    return sorted_list


def write_freq_dist_to_workbook(
    f_dist: FreqDist, path: str = "../gradu/material/final_results.xlsx"
):
    """Writes the frequency distribution of keywords on the the worksheet"""
    data_column = "B"
    data_row = 3
    wb = load_workbook(path)
    if not "Keywords" in wb:
        wb.create_sheet(title="Keywords", index=2)
    sheet = wb["Keywords"]
    for word, count in sorted(f_dist.items(), key=lambda x: x[1], reverse=True):
        sheet[data_column + str(data_row)] = f"{word}/{count}"
        data_row += 1
    wb.save(path)


def finalize_categories(path: str = "../gradu/material/final_results.xlsx"):
    """Updates the category numbers and writes them to the workbook"""
    mapping = {
        1: 1,
        8: 2,
        15: 3,
        13: 4,
        17: 5,
        23: 6,
        3: 7,
        22: 8,
        9: 9,
        20: 10,
        24: 11,
        26: 12,
        4: 13,
        6: 14,
        16: 15,
        5: 16,
        2: 17,
        14: 18,
        7: 19,
        10: 20,
        11: 21,
        18: 22,
    }

    old_cats_column = "P"
    new_cats_column = "Q"

    wb = load_workbook(path)
    sheet = wb["Accepted Papers"]
    for i in range(3, 95):
        row = str(i)
        cat_string = sheet[old_cats_column + row].value
        cats = parse_cat_numbers(cat_string)
        new_cats = (str(mapping[x]) for x in cats)
        new_cats_str = ", ".join(new_cats)
        sheet[new_cats_column + row] = new_cats_str
    wb.save(path)


def save_wieringa_plot(papers: pd.DataFrame):
    """"Draws the Wieringa classifications as a bar plot and saves it to a file"""
    data = []
    for c in (x[0] for x in W_CLASSES):
        data.append(sum(papers[c]))

    # ER, VR, SP, PP, OP, PEP
    types = [x[1] for x in W_CLASSES]
    print(types)
    y_pos = np.arange(len(types))
    plt.barh(y_pos, data)
    plt.yticks(y_pos, types)
    plt.tight_layout()
    plt.savefig("../gradu/material/data/wierienga_bar.png", dpi=400)


def get_wieringa_topic_bubble_data(papers: List[Paper]):
    """"Writes the data for Wieringa classifications and topics bubble chart"""
    data = {}
    cat_dict = {k: v for k, v in CATS}
    w_classes_dict = {k: v for k, v in W_CLASSES}
    for paper in papers:
        for cat in paper.categories:
            for w in paper.w_classes:
                point = (cat_dict[cat], w_classes_dict[w])
                if point in data:
                    data[point] += 1
                else:
                    data[point] = 1
    data_lines = (
        f"{key[0]},{key[1]},{count}\n"
        for key, count in sorted(data.items(), key=lambda x: x[0])
    )
    with open("../gradu/material/data/class_topic_bubble.dat", "w") as f:
        f.write("Category,Class,Count\n")
        f.writelines(data_lines)


def save_forum_pie(papers):
    """Creates a pie chart of publication venues and saves it to a file"""
    venues = {}
    for paper in papers:
        if paper.venue not in venues:
            venues[paper.venue] = 1
        else:
            venues[paper.venue] += 1

    labels = []
    values = []
    for k, v in venues.items():
        freq = v / len(papers)
        values.append(freq)
        pct = f"{freq*100:.2f}%"
        labels.append(f" {k} ({pct})")

    plt.pie(values, labels=labels)
    plt.savefig("../gradu/material/data/forum_pie.png", dpi=400)


def save_yearly_publications(papers: List[Paper]):
    """Creates a chart of publications per year per venue and saves it to a file"""
    years = [2015, 2016, 2017, 2018, 2019]
    venues = ["JAIR", "IJCAI", "AIJ", "JAGI", "ICAGI"]

    ys = {v: {year: 0 for year in years} for v in venues}
    for paper in papers:
        ys[paper.venue][paper.year] += 1

    y_jair = list(ys["JAIR"].values())
    y_ijcai = list(ys["IJCAI"].values())
    y_aij = list(ys["AIJ"].values())
    y_jagi = list(ys["JAGI"].values())
    y_icagi = list(ys["ICAGI"].values())

    y_all = np.array([y_jair, y_ijcai, y_aij, y_jagi, y_icagi])

    for i in range(len(y_all)):
        plt.bar(years, y_all[i], bottom=np.sum(y_all[:i], axis=0))

    plt.legend(venues)
    plt.savefig("../gradu/material/data/yearly_publications.png", dpi=600)


def save_topic_frequencies_by_year(papers: List[Paper]):
    """Creates horizontal bar plot of topics with their yearly stacked frequencies"""
    years = [2015, 2016, 2017, 2018, 2019]

    category_names = [x[0] for x in CATS]
    ys = {year: {c: 0 for c in category_names} for year in years}

    for paper in papers:
        for cat in paper.categories:
            ys[paper.year][cat] += 1

    y2015 = list(ys[2015].values())
    y2016 = list(ys[2016].values())
    y2017 = list(ys[2017].values())
    y2018 = list(ys[2018].values())
    y2019 = list(ys[2019].values())

    y_all = np.array([y2015, y2016, y2017, y2018, y2019])

    fig, ax = plt.subplots()

    for i in range(len(y_all)):
        ax.barh(category_names, y_all[i], left=np.sum(y_all[:i], axis=0))

    ax.set_yticks(range(1, len(category_names) + 1))
    ax.set_yticklabels([x[1] for x in CATS])

    plt.legend(years)
    plt.tight_layout()  # TODO: Figure size could be customized instead of this

    plt.savefig("../gradu/material/data/topic_frequencies_by_year.png", dpi=400)


def save_topic_frequencies():
    cat_dict = calculate_category_count()
    xs = []
    ys = []
    for k, v in sorted(cat_dict.items(), key=lambda x: x[0]):
        xs.append(k)
        ys.append(v)

    fig, ax = plt.subplots()
    # fig.set_figheight(5)
    # fig.set_figwidth(10)
    plt.tight_layout()
    ax.barh(xs, ys)
    ax.set_yticks(xs)
    ax.set_yticklabels([x[1] for x in CATS])
    plt.savefig(
        "../gradu/material/data/topic_frequencies.png", dpi=400, bbox_inches="tight"
    )


def save_topic_heatmap(papers: List[Paper]):
    """Creates a square matrix heatmap from the topic categories and saves it."""
    category_names = [x[1] for x in CATS]
    matrix = np.zeros(shape=(23, 23))

    # Get all sum of all combinations
    for paper in papers:
        combs = ((a, b) for a in paper.categories for b in paper.categories)
        for comb in combs:
            matrix[comb] += 1

    data = matrix[1:, 1:]

    # Plot the heatmap figure
    # Here, Gridspec is used to define the grid where the plots are drawn.
    # The colorbar is a separate subplot so it can be more easily be moved
    fig = plt.figure(figsize=(9, 9))
    gs = gridspec.GridSpec(nrows=2, ncols=8, height_ratios=[0.05, 1])
    ax = fig.add_subplot(gs[1, :])
    cax = fig.add_subplot(gs[0, 1:7])

    # Create image from data matrix
    im = ax.imshow(data, cmap="hot")

    # Create colorbar and its title
    cbar = fig.colorbar(im, cax=cax, shrink=0.5, orientation="horizontal")
    cbar.ax.set_title("No. of articles")

    # Set ticks and labels
    ax.set_xticks(np.arange(len(category_names)))
    ax.set_yticks(np.arange(len(category_names)))
    ax.set_xticklabels(category_names, fontsize=9)
    ax.set_yticklabels(category_names, fontsize=9)

    plt.setp(ax.get_xticklabels(), rotation=30, ha="right", rotation_mode="anchor")

    # Text annotation colors and their threshold, alignment
    textcolors = ("black", "white")
    kw = dict(horizontalalignment="center", verticalalignment="center")
    threshold = im.norm(data.max()) / 2.0

    # Loop over data dimensions and create text annotations for cells.
    texts = []
    for i in range(len(category_names)):
        for j in range(len(category_names)):
            kw.update(color=textcolors[int(im.norm(data[i, j]) < threshold)])
            txt = f"{data[i, j]:.0f}" if data[i, j] != 0 else ""
            text = im.axes.text(j, i, txt, **kw)
            texts.append(text)

    plt.tight_layout()
    plt.savefig(
        "../gradu/material/data/topic_heatmap_no_zeroes.png",
        dpi=400,
        bbox_inches="tight",
    )


def calculate_country_frequencies(papers: List[Paper]) -> Dict[str, int]:
    """ Calculates frequencies of the countries in papers, and returns results as a dict"""
    results = {}
    for paper in papers:
        for country in paper.countries:
            if country in results:
                results[country] += 1
            else:
                results[country] = 1
    return results


def find_country_data(
    papers: List[Paper],
    path: str = "../gradu/material/data/country_data.csv",
    count_desc: bool = True,
):
    """Find location data of countries and save all necessary data to csv file"""
    country_dict = calculate_country_frequencies(papers)
    del country_dict["none"]
    app = Nominatim(user_agent="ai-thesis-mapping")

    header = "country,count,lat,lon".split(sep=",")
    rows = []
    for k, v in country_dict.items():
        location = app.geocode(k)
        country = k.title() if len(k) > 3 else k.upper()
        row = [country, v, location.latitude, location.longitude]
        rows.append(row)

    rows = sorted(rows, key=lambda x: x[1], reverse=count_desc)

    with open(path, "w", newline="", encoding="utf-8") as f:
        csvwriter = csv.writer(f)
        csvwriter.writerow(header)
        csvwriter.writerows(rows)
        # The non-country paper is added back
        f.write("None (Microsoft),1,0,0")


def draw_map(papers: List[Paper]):
    """" Draw a bubble map with countries' papers showing"""

    sizeref = 2 * 36 / (70 ** 2)
    df = pd.read_csv("../gradu/material/data/country_data.csv")
    df["text"] = df["count"].astype(str)
    fig = go.Figure(
        data=go.Scattergeo(
            lon=df["lon"],
            lat=df["lat"],
            text=df["text"],
            mode="markers+text",
            marker=dict(
                color=df["count"],
                colorscale="Delta",
                size=df["count"],
                sizemode="area",
                sizeref=sizeref,
                sizemin=7,
            ),
            textfont=dict(family="sans serif", color="white"),
        )
    )
    fig.update_geos(projection_type="natural earth")
    fig.show()


def main():
    # Initial keyword extraction is done and written to workbook
    # papers = load_papers()
    # all_keywords = []
    # for p in papers:
    #     all_keywords += p.keywords
    # fdist = FreqDist(all_keywords)
    # write_freq_dist_to_workbook(fdist)

    # Counts the similar keyword instances on a row and writes it to workbook
    # count_keywords()

    # Category distribution is written to the workbook
    # c_dist = calculate_category_distribution()
    # write_category_dist_to_workbook(c_dist)

    # Amount of articles in categories is printed
    # cat_count = calculate_category_count()
    # print(cat_count)

    # Updates category numbers and writes them into workbook
    # finalize_categories()

    papers = load_papers()
    # papers_df = load_papers_df()

    # pp = pprint.PrettyPrinter()
    # pp.pprint(papers)

    # === Drawing the graphs ===

    # save_wieringa_plot(papers)
    # get_wieringa_topic_bubble_data(load_papers())

    # Forums pie chart
    # save_forum_pie(papers)

    # Publication years
    # save_yearly_publications(papers)

    # Topic frequensies
    # save_topic_frequencies_by_year(papers)

    # Topic heatmap/square matrix
    # save_topic_heatmap(papers)

    # Calculate countries
    # print(calculate_country_frequencies(papers))

    # Find and save country data to csv file
    # find_country_data(papers, "../gradu/material/data/country_data_desc.csv")

    # Draw the geographic map with article counts per country
    # draw_map(papers)

    # Create csv of accepted papers
    create_article_csv(papers)

    # Shows the papers that are included in given categories
    if 1 < len(sys.argv):
        cat_numbers = sys.argv[1:]
        cat_numbers = list(map(int, cat_numbers))
        papers = get_category_papers(cat_numbers)
        pp = pprint.PrettyPrinter()
        pp.pprint(papers)

    # Shows the papers that are included in given countries
    # if 1 < len(sys.argv):
    #     countries = sys.argv[1:]
    #     papers = get_country_papers(countries)
    #     # authors = [x.authors for x in papers]
    #     pp = pprint.PrettyPrinter()
    #     pp.pprint(papers)


if __name__ == "__main__":
    main()
